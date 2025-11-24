import os
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.pyplot as plt

# ------------------------ CONFIGURAÇÃO DE LOGGING ------------------------

LOG_FILE = "coin.log"

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,  # pode trocar para DEBUG se quiser mais detalhe
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    encoding="utf-8"
)

logger = logging.getLogger("coin")  # logger principal do sistema


# ------------------------ FUNÇÕES PRINCIPAIS ------------------------


def criar_planilha():
    arquivo = "Controle_Financeiro.xlsx"

    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        if "Transacoes" in wb.sheetnames:
            ws = wb["Transacoes"]
            logger.info("Planilha existente carregada: %s (aba 'Transacoes')", arquivo)
        else:
            ws = wb.active
            ws.title = "Transacoes"
            if ws.max_row == 1 and all(c.value is None for c in ws[1]):
                ws.append([
                    "ID",
                    "Valor",
                    "Tipo",
                    "Categoria",
                    "Descricao",
                    "Dia",
                    "Mes",
                    "Ano"
                ])
                wb.save(arquivo)
            logger.info("Planilha existente ajustada com aba 'Transacoes' e cabeçalho.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transacoes"

        ws.append([
            "ID",
            "Valor",
            "Tipo",
            "Categoria",
            "Descricao",
            "Dia",
            "Mes",
            "Ano"
        ])

        wb.save(arquivo)
        logger.info("Nova planilha criada: %s (aba 'Transacoes')", arquivo)

    return wb, ws


def adicionar_transa(wb, ws):
    """
    Versão em modo texto (terminal). No seu fluxo atual com Tkinter,
    você está usando a versão gráfica (gui_adicionar_transacao),
    mas mantive esta função por compatibilidade.
    """
    logger.info("Iniciando cadastro de transações (modo terminal).")
    print("----------------------------Cadastro de transações-----------------------------")

    max_id = 0
    for linha in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        cell_id = linha[0]
        if cell_id is None:
            continue
        try:
            iid = int(cell_id)
            if iid > max_id:
                max_id = iid
        except:
            continue

    id_transacao = max_id + 1

    while True:

        print(f"\nRegistrando transação de ID {id_transacao}...\n")

        # Cadastrando o Valor
        while True:
            valor_numero = input("Digite o valor da transação: ")
            valor_numero = valor_numero.replace(",", ".")
            try:
                valor = float(valor_numero)
                if valor > 0:
                    break
                print("O valor deve ser maior que zero.")
            except:
                print("Valor inválido. Tente novamente.")

        # Cadastrando o Tipo
        while True:
            tipo_transa = input("Digite o tipo da transação (1 - entrada, 2 - saída): ")
            if tipo_transa == "1":
                tipo = "entrada"
                break
            elif tipo_transa == "2":
                tipo = "saida"
                break
            else:
                print("Opção inválida. Digite 1 ou 2.")

        # Cadastrando Categoria
        while True:
            categoria = input("Categoria (lazer, alimento, trabalho, estudos): ").strip().lower()
            if categoria in ["lazer", "alimento", "trabalho", "estudos"]:
                break
            print("Categoria inválida.")

        # Cadastrando Descrição
        while True:
            descricao = input("Descrição: ").strip()
            if descricao != "":
                break
            print("Descrição não pode ser vazia.")

        # Cadastrando o Dia
        while True:
            dia_numero = input("Dia (1 a 30): ")
            if dia_numero.isdigit() and 1 <= int(dia_numero) <= 30:
                dia = int(dia_numero)
                break
            print("Dia inválido.")

        # Cadastrando Mês
        while True:
            mes_numero = input("Mês (1 a 12): ")
            if mes_numero.isdigit() and 1 <= int(mes_numero) <= 12:
                mes = int(mes_numero)
                break
            print("Mês inválido.")

        # Cadastrando Ano
        while True:
            ano_numero = input("Ano (1 a 9999): ")
            if ano_numero.isdigit() and 1 <= int(ano_numero) <= 9999:
                ano = int(ano_numero)
                break
            print("Ano inválido.")

        ws.append([id_transacao, valor, tipo, categoria, descricao, dia, mes, ano])
        wb.save("Controle_Financeiro.xlsx")
        logger.info(
            "Transação adicionada (terminal): ID=%s, valor=%.2f, tipo=%s, categoria=%s, data=%02d/%02d/%04d",
            id_transacao, valor, tipo, categoria, dia, mes, ano
        )
        print("Transação salva com sucesso!")

        id_transacao += 1

        continuar = input("\nDeseja registrar outra transação? (s/n): ").lower().strip()
        if continuar != "s":
            print("Encerrando cadastro.")
            break

    print("Arquivo atualizado: Controle_Financeiro.xlsx")


#------------------------------Opção 2------------------------------


def remover_transa(wb, ws, id_remover):

    for indice_linha, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cell_id = row[0].value
        if cell_id == id_remover:
            valor = row[1].value
            tipo = row[2].value
            categoria = row[3].value
            descricao = row[4].value
            dia = row[5].value
            mes = row[6].value
            ano = row[7].value

            detalhes = (
                f"ID: {cell_id}\n"
                f"Valor: {valor}\n"
                f"Tipo: {tipo}\n"
                f"Categoria: {categoria}\n"
                f"Data: {dia}/{mes}/{ano}\n"
                f"Descrição: {descricao}"
            )

            ws.delete_rows(indice_linha, 1)
            wb.save("Controle_Financeiro.xlsx")
            logger.info("Transação removida: ID=%s, valor=%s, tipo=%s, categoria=%s, data=%s/%s/%s",
                        cell_id, valor, tipo, categoria, dia, mes, ano)
            return True, detalhes

    logger.warning("Tentativa de remover transação inexistente. ID=%s", id_remover)
    return False, "Nenhuma transação com esse ID foi encontrada."


#------------------------------Opção 3-------------------------------


def listar_por_categoria(ws, categoria_busca):
    categoria_busca = str(categoria_busca).strip().lower()

    if categoria_busca not in ["lazer", "alimento", "trabalho", "estudos"]:
        msg = "Categoria inválida. Use: lazer, alimento, trabalho ou estudos."
        logger.warning("Categoria inválida em listar_por_categoria: %s", categoria_busca)
        print(msg)
        return msg

    linhas_resultado = []
    linhas_resultado.append(f"Transações da categoria: {categoria_busca}")
    linhas_resultado.append("ID | Valor | Tipo | Categoria | Data       | Descrição")
    linhas_resultado.append("-----------------------------------------------------------")

    print("\n----------------------------Listar por categoria-----------------------------")
    print(f"Categoria buscada: {categoria_busca}")
    print("ID | Valor | Tipo | Categoria | Data       | Descrição")
    print("-----------------------------------------------------------")

    encontrou = False
    total_saidas = 0.0
    qtd_saidas = 0

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha:
            continue

        tid, valor, tipo, categoria, descricao, dia, mes, ano = linha

        if categoria is None:
            continue

        if str(categoria).strip().lower() == categoria_busca:
            encontrou = True
            data_str = f"{int(dia):02d}/{int(mes):02d}/{int(ano)}"
            texto = f"{tid} | {valor} | {tipo} | {categoria} | {data_str} | {descricao}"
            print(texto)
            linhas_resultado.append(texto)

            tipo_norm = str(tipo).strip().lower()
            tipo_norm = tipo_norm.replace("í", "i").replace("á", "a")

            if "saida" in tipo_norm:
                total_saidas += float(valor)
                qtd_saidas += 1

    if not encontrou:
        msg = "Nenhuma transação encontrada para essa categoria."
        logger.info("listar_por_categoria sem resultados. Categoria=%s", categoria_busca)
        print(msg)
        return msg

    linhas_resultado.append("-----------------------------------------------------------")
    print("-----------------------------------------------------------")
    linhas_resultado.append(f"Total de GASTOS (saídas) na categoria '{categoria_busca}': R$ {total_saidas:.2f}")
    print(f"Total de GASTOS (saídas) na categoria '{categoria_busca}': R$ {total_saidas:.2f}")

    if qtd_saidas > 0:
        media_gastos = total_saidas / qtd_saidas
        linhas_resultado.append(f"MÉDIA de gasto por transação nessa categoria: R$ {media_gastos:.2f}")
        print(f"MÉDIA de gasto por transação nessa categoria: R$ {media_gastos:.2f}")
    else:
        linhas_resultado.append("Não houve saídas (gastos) nessa categoria, portanto não há média de gastos.")
        print("Não houve saídas (gastos) nessa categoria, portanto não há média de gastos.")

    logger.info("listar_por_categoria: categoria=%s, total_saidas=%.2f, qtd_saidas=%d",
                categoria_busca, total_saidas, qtd_saidas)

    return "\n".join(linhas_resultado)


#______________________________Opções Gráficas________________________


def grafico_pizza_categorias(ws):
    totais_por_categoria = {}

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha:
            continue

        tid, valor, tipo, categoria, descricao, dia, mes, ano = linha

        if valor is None or categoria is None or tipo is None:
            continue

        cat_norm = str(categoria).strip().lower()
        tipo_norm = str(tipo).strip().lower()
        tipo_norm = tipo_norm.replace("í", "i").replace("á", "a")

        if "saida" in tipo_norm:
            try:
                v = float(valor)
            except:
                continue
            if v <= 0:
                continue
            totais_por_categoria[cat_norm] = totais_por_categoria.get(cat_norm, 0) + v

    if not totais_por_categoria:
        msg = "Não há saídas registradas para gerar o gráfico de pizza."
        logger.info("grafico_pizza_categorias: nenhum dado de saída encontrado.")
        return msg

    labels = list(totais_por_categoria.keys())
    valores = list(totais_por_categoria.values())

    logger.info("Gerando gráfico de pizza. Categorias=%s", labels)

    plt.figure()
    plt.pie(valores, labels=labels, autopct="%1.1f%%")
    plt.title("Proporção de gastos por categoria")
    plt.tight_layout()
    plt.show()

    return "Gráfico de pizza exibido com sucesso."


def grafico_saldo_acumulado(ws):
    """
    Cria um gráfico de linha mostrando o saldo acumulado ao longo do tempo.
    Usa o índice das transações no eixo X e mostra as datas como rótulos.
    """
    dados = []

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha:
            continue

        tid, valor, tipo, categoria, descricao, dia, mes, ano = linha

        if valor is None or tipo is None or dia is None or mes is None or ano is None:
            continue

        try:
            ano_int = int(ano)
            mes_int = int(mes)
            dia_int = int(dia)

            if not (1 <= ano_int <= 9999):
                continue

            data = datetime(ano_int, mes_int, dia_int)
            v = float(valor)
        except Exception:
            continue

        tipo_norm = str(tipo).strip().lower()
        tipo_norm = tipo_norm.replace("í", "i").replace("á", "a")

        if "entrada" in tipo_norm:
            tipo_final = "entrada"
        elif "saida" in tipo_norm:
            tipo_final = "saida"
        else:
            continue

        dados.append((data, tipo_final, v))

    if not dados:
        logger.info("grafico_saldo_acumulado: nenhum dado de transação para plotar.")
        return "Não há transações suficientes para gerar o gráfico de saldo."

    dados.sort(key=lambda x: x[0])

    saldos = []
    datas_str = []
    saldo_atual = 0.0

    for data, tipo_final, v in dados:
        if tipo_final == "entrada":
            saldo_atual += v
        elif tipo_final == "saida":
            saldo_atual -= v

        saldos.append(saldo_atual)
        datas_str.append(data.strftime("%d/%m/%Y"))

    if not saldos:
        logger.warning("grafico_saldo_acumulado: lista de saldos vazia após processamento.")
        return "Não foi possível gerar o gráfico: nenhuma transação válida encontrada."

    x = list(range(len(saldos)))

    logger.info("Gerando gráfico de saldo acumulado. Total de pontos=%d", len(saldos))

    plt.figure()
    plt.plot(x, saldos, marker="o")
    plt.xlabel("Transações (em ordem cronológica)")
    plt.ylabel("Saldo acumulado (R$)")
    plt.title("Evolução do saldo acumulado ao longo do tempo")
    plt.xticks(x, datas_str, rotation=45, ha="right")
    plt.tight_layout()
    plt.show()

    return "Gráfico de linha exibido com sucesso."


#------------------------------Opção 4-------------------------------


def listar_por_periodo(ws, data_inicial, data_final):
    if data_inicial > data_final:
        msg = "Período inválido: a data inicial é maior que a final."
        logger.warning("listar_por_periodo com período inválido: %s > %s", data_inicial, data_final)
        print(msg)
        return msg

    linhas_resultado = []
    linhas_resultado.append(f"Transações de {data_inicial.strftime('%d/%m/%Y')} até {data_final.strftime('%d/%m/%Y')}")
    linhas_resultado.append("ID | Valor | Tipo | Categoria | Data       | Descrição")
    linhas_resultado.append("-----------------------------------------------------------")

    print("\nTransações dentro do período:")
    print("ID | Valor | Tipo | Categoria | Data       | Descrição")
    print("-----------------------------------------------------------")

    encontrou = False
    total_saidas = 0.0
    qtd_saidas = 0

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha:
            continue

        tid, valor, tipo, categoria, descricao, dia, mes, ano = linha

        if dia is None or mes is None or ano is None or valor is None:
            continue

        try:
            data_transacao = datetime(int(ano), int(mes), int(dia))
        except:
            continue

        if data_inicial <= data_transacao <= data_final:
            encontrou = True
            data_str = data_transacao.strftime("%d/%m/%Y")
            texto = f"{tid} | {valor} | {tipo} | {categoria} | {data_str} | {descricao}"
            print(texto)
            linhas_resultado.append(texto)

            tipo_norm = str(tipo).strip().lower()
            tipo_norm = tipo_norm.replace("í", "i").replace("á", "a")
            if "saida" in tipo_norm:
                total_saidas += float(valor)
                qtd_saidas += 1

    print("-----------------------------------------------------------")
    linhas_resultado.append("-----------------------------------------------------------")

    if not encontrou:
        msg = "Nenhuma transação encontrada nesse período."
        logger.info("listar_por_periodo sem resultados. Periodo %s a %s", data_inicial, data_final)
        print(msg)
        return msg

    resumo = f"Total de GASTOS (saídas) no período: R$ {total_saidas:.2f}"
    print(resumo)
    linhas_resultado.append(resumo)

    if qtd_saidas > 0:
        media_gastos = total_saidas / qtd_saidas
        resumo2 = f"MÉDIA de gasto por transação no período: R$ {media_gastos:.2f}"
        print(resumo2)
        linhas_resultado.append(resumo2)
    else:
        resumo2 = "Não houve saídas (gastos) no período, portanto não há média de gastos."
        print(resumo2)
        linhas_resultado.append(resumo2)

    logger.info(
        "listar_por_periodo: periodo=%s a %s, total_saidas=%.2f, qtd_saidas=%d",
        data_inicial, data_final, total_saidas, qtd_saidas
    )

    return "\n".join(linhas_resultado)


#------------------------------Opção 5-------------------------------


def calcular_saldo_periodo(ws, data_inicial, data_final):
    if data_inicial > data_final:
        msg = "Período inválido: a data inicial é maior que a final."
        logger.warning("calcular_saldo_periodo com período inválido: %s > %s", data_inicial, data_final)
        print(msg)
        return msg

    saldo = 0.0
    total_entradas = 0.0
    total_saidas = 0.0
    encontrou = False

    saldo_por_mes = {}

    for linha in ws.iter_rows(min_row=2, values_only=True):

        if not linha:
            continue

        tid, valor, tipo, categoria, descricao, dia, mes, ano = linha

        if dia is None or mes is None or ano is None or valor is None or tipo is None:
            continue

        try:
            ano_int = int(ano)
            mes_int = int(mes)
            dia_int = int(dia)
            if not (1 <= ano_int <= 9999):
                continue
            data_transacao = datetime(ano_int, mes_int, dia_int)
        except:
            continue

        if data_inicial <= data_transacao <= data_final:
            encontrou = True
            valor_f = float(valor)

            chave_mes = (ano_int, mes_int)
            if chave_mes not in saldo_por_mes:
                saldo_por_mes[chave_mes] = 0.0

            tipo_norm = str(tipo).strip().lower()
            tipo_norm = tipo_norm.replace("í", "i").replace("á", "a")

            if "entrada" in tipo_norm:
                total_entradas += valor_f
                saldo += valor_f
                saldo_por_mes[chave_mes] += valor_f

            elif "saida" in tipo_norm:
                total_saidas += valor_f
                saldo -= valor_f
                saldo_por_mes[chave_mes] -= valor_f

    linhas_resultado = []
    linhas_resultado.append("================== RESULTADO DO PERÍODO ==================")
    cabecalho = f"Período: {data_inicial.strftime('%d/%m/%Y')}  até  {data_final.strftime('%d/%m/%Y')}"
    linhas_resultado.append(cabecalho)
    linhas_resultado.append("----------------------------------------------------------")

    print("\n================== RESULTADO DO PERÍODO ==================")
    print(cabecalho)
    print("----------------------------------------------------------")

    if not encontrou:
        msg = "Nenhuma transação encontrada nesse período."
        logger.info("calcular_saldo_periodo sem resultados. Periodo %s a %s", data_inicial, data_final)
        print(msg)
        linhas_resultado.append(msg)
        return "\n".join(linhas_resultado)

    txt_ent = f"Total de ENTRADAS: R$ {total_entradas:.2f}"
    txt_sai = f"Total de SAÍDAS..: R$ {total_saidas:.2f}"
    txt_saldo = f"SALDO NO PERÍODO.: R$ {saldo:.2f}"

    print(txt_ent)
    print(txt_sai)
    print("----------------------------------------------------------")
    print(txt_saldo)

    linhas_resultado.append(txt_ent)
    linhas_resultado.append(txt_sai)
    linhas_resultado.append("----------------------------------------------------------")
    linhas_resultado.append(txt_saldo)

    linhas_resultado.append("\n================== SALDO POR MÊS ==================")
    print("\n================== SALDO POR MÊS ==================")

    chaves_ordenadas = sorted(saldo_por_mes.keys())

    for (ano_m, mes_m) in chaves_ordenadas:
        saldo_mes = saldo_por_mes[(ano_m, mes_m)]
        linha_mes = f"Mês {mes_m:02d}/{ano_m}: R$ {saldo_mes:.2f}"
        print(linha_mes)
        linhas_resultado.append(linha_mes)

    logger.info(
        "calcular_saldo_periodo: periodo=%s a %s, entradas=%.2f, saidas=%.2f, saldo=%.2f",
        data_inicial, data_final, total_entradas, total_saidas, saldo
    )

    return "\n".join(linhas_resultado)
